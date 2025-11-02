import pandas as pd
from io import BytesIO
from typing import Dict, Optional, Union
from app.utils.logs import logger


code_bank_duplicate = '471000'

def generate_bank_statement_excel(
    data: Dict, 
    output_path: Optional[str] = None
) -> Union[str, BytesIO]:
    """
    Génère un fichier Excel à partir des données de relevé bancaire.
    """
    logger.info("Génération du fichier Excel pour les données du relevé bancaire")
    
    # Crée un objet BytesIO ou utilise le chemin de sortie
    if output_path:
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        return_path = True
    else:
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        return_path = False
    
    # Crée la feuille de synthèse du compte
    summary_data = {
        'Information': [
            'Nom de la banque', 
            'Numéro de compte', 
            'Titulaire du compte', 
            'Date de début',
            'Date de fin', 
            'Solde initial', 
            'Solde final', 
            'Devise'
        ],
        'Value': [
            data.get('bank_name', ''),
            data.get('account_number', ''),
            data.get('account_holder', ''),
            data.get('starting_date', ''),
            data.get('closing_date', ''),
            data.get('starting_balance', 0),
            data.get('closing_balance', 0),
            data.get('currency', '')
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Synthèse du compte', index=False)
    
    # Crée la feuille des transactions
    transactions = data.get('transactions', [])
    if transactions:
        transactions_df = pd.DataFrame(transactions)
    else:
        transactions_df = pd.DataFrame(columns=[
            'date', 'document_number', 'bank_code', 'account_number', 
            'description', 'debit', 'credit'
        ])

    # Incrémenter les numéros de pièce de façon robuste
    if not transactions_df.empty and 'document_number' in transactions_df.columns:
        # Convertir en numérique (int). Les valeurs non numériques deviennent 0.
        transactions_df['document_number'] = pd.to_numeric(
            transactions_df['document_number'], errors='coerce'
        ).fillna(0).astype(int)

        # Ajouter un offset séquentiel (0, 1, 2, ...) pour chaque ligne
        transactions_df['document_number'] = (
            transactions_df['document_number'] + pd.RangeIndex(start=0, stop=len(transactions_df))
        )


    # Dupliquer les lignes en changeant certaines informations
    if not transactions_df.empty:
        rows = []
        # itérate uniquement sur les lignes originales
        for _, orig in transactions_df.iterrows():
            original_row = orig.copy()
            duplicated_row = orig.copy()
            # modifier les champs du duplicata
            duplicated_row['bank_code'] = code_bank_duplicate
            duplicated_row['debit'] = orig.get('credit', 0)
            duplicated_row['credit'] = orig.get('debit', 0)
            # ajouter d'abord l'original puis son duplicata
            rows.append(original_row.to_dict())
            rows.append(duplicated_row.to_dict())
        # reconstruire le DataFrame dans l'ordre voulu
        transactions_df = pd.DataFrame(rows, columns=transactions_df.columns)
    
    # Renomme les colonnes pour un meilleur affichage
    rename_map = {
        'date': 'Date',
        'document_number': 'Numéro de pièce',
        'bank_code': 'Code banque',
        'account_number': 'Numéro de compte',
        'description': 'Description',
        'debit': 'Débit',
        'credit': 'Crédit'
    }
    transactions_df = transactions_df.rename(columns={k: v for k, v in rename_map.items() if k in transactions_df.columns})
    
    # Ajouter les lignes de totaux
    if not transactions_df.empty and 'Débit' in transactions_df.columns and 'Crédit' in transactions_df.columns:
        # Calculer les totaux
        total_debit = transactions_df['Débit'].sum()
        total_credit = transactions_df['Crédit'].sum()
        
        # Créer les lignes de totaux
        empty_row = pd.Series([''] * len(transactions_df.columns), index=transactions_df.columns)
        
        total_debit_row = empty_row.copy()
        total_debit_row['Date'] = 'Total'
        total_debit_row['Débit'] = total_debit
        
        total_credit_row = empty_row.copy()
        total_credit_row['Crédit'] = total_credit
        
        # Ajouter les lignes au DataFrame
        transactions_df = pd.concat([
            transactions_df,
            pd.DataFrame([total_debit_row]),
            pd.DataFrame([total_credit_row])
        ], ignore_index=True)
    
    transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
    
    # Formate les colonnes avec openpyxl
    workbook = writer.book
    
    # Formate la feuille de synthèse
    summary_sheet = writer.sheets['Synthèse du compte']
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 30
    
    # Formate la feuille des transactions
    transactions_sheet = writer.sheets['Transactions']
    col_widths = {
        'A': 15,  # Date
        'B': 20,  # Numéro de document
        'C': 15,  # Code banque
        'D': 20,  # Numéro de compte
        'E': 70,  # Description
        'F': 15,  # Débit
        'G': 15   # Crédit
    }
    for col, width in col_widths.items():
        if col in transactions_sheet.column_dimensions:
            transactions_sheet.column_dimensions[col].width = width
    
    # Formatage spécial pour les lignes de totaux
    if not transactions_df.empty:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Trouver les lignes de totaux (les 2 dernières lignes non-vides)
        last_row = len(transactions_df) + 1  # +1 car on compte l'en-tête
        
        # Style pour les totaux (gras et fond gris clair)
        bold_font = Font(bold=True)
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Appliquer le style aux deux dernières lignes (totaux)
        for row in range(last_row - 1, last_row + 1):  # 2 dernières lignes
            for col in range(1, len(transactions_df.columns) + 1):
                cell = transactions_sheet.cell(row=row, column=col)
                cell.font = bold_font
                cell.fill = gray_fill
        
        # Fusionner les cellules de la colonne Date (colonne A) pour le titre "Total"
        transactions_sheet.merge_cells(f'A{last_row - 1}:A{last_row}')
        
        # Appliquer l'alignement centré à la cellule fusionnée
        merged_cell = transactions_sheet.cell(row=last_row - 1, column=1)
        merged_cell.alignment = center_alignment
    
    # Sauvegarde le classeur
    logger.info("Sauvegarde du classeur Excel")
    writer.close()
    
    if return_path:
        logger.info(f"Fichier Excel sauvegardé à : {output_path}")
        return output_path
    else:
        logger.info("Fichier Excel généré en mémoire")
        output.seek(0)
        return output

def save_excel_to_file(excel_data: BytesIO, filepath: str) -> str:
    """
    Sauvegarde un objet Excel BytesIO dans un fichier
    
    Args:
        excel_data: Objet BytesIO contenant les données Excel
        filepath: Chemin où le fichier doit être sauvegardé
        
    Returns:
        Le chemin du fichier sauvegardé
    """
    logger.info(f"Sauvegarde des données Excel dans le fichier : {filepath}")
    with open(filepath, 'wb') as f:
        f.write(excel_data.getvalue())
    logger.info(f"Fichier Excel sauvegardé avec succès à : {filepath}")
    return filepath
